from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series

dict_anos = {}

# 1 - Importando DEspesas 
arquivo1 = load_workbook(filename='files/Despesas.xlsx')
ws1 = arquivo1['Despesas']
max_linhas = ws1.max_row

for i in range(2, max_linhas+1):
    # print(ws1['A%' %i].value)
    # print(ws1['B%' %i].value)
    dict_anos[ws1['A%s' %i].value] = {'Despesas':ws1['B%s' %i].value, 'Receitas':0}

# print(dict_anos)

# 2 - Importando Receita
arquivo2 = load_workbook(filename='files/Receitas.xlsx')
ws2 = arquivo2['Receitas']
max_linhas = ws2.max_row

for i in range(2, max_linhas+1):
    dict_anos[ws2['A%s' %i].value]['Receitas'] = ws2['B%s' %i].value
    
print(dict_anos)

# 3 - Criando a planilha

wb = Workbook()
ws = wb.active

ws['A1'] = 'Ano'
ws['B1'] = 'Despesas'
ws['C1'] = 'Receitas'

i = 2
for key, value in dict_anos.items():
    # print(key)
    # print(value)
    ws['A%s' %i] = key
    ws['B%s' %i] = value['Despesas']
    ws['C%s' %i] = value['Receitas']
    i += 1
    
    chart1 = BarChart()
    chart1.type = 'col'
    chart1.style =  12
    chart1.title = 'Receitas X Despesas por Ano'
    chart1.y_axis.title = 'R$'
    chart1.x_axis.title = 'Ano'
    
    data = Reference(
        ws,
        min_col=2,
        max_col=3,
        min_row=1,
        max_row=i
    )
    
    anos = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_row=i
    )
    
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(anos)
    chart1.shape = 4
    ws.add_chart(chart1, 'A%s' %(i+2))
    
wb.save(filename='files/demonstrativo.xlsx')